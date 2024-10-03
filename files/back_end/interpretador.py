import os
import sqlite3
import openpyxl
import datetime
from PyQt5.QtWidgets import QMessageBox
import analista_dados.files.front_end.loading_interpretador as loading_interpretador

global main_window

def dicionario_xlsx(caminho_arquivo, supervisor, unidade, parent):
    """Processa um arquivo .xlsx e insere os dados no banco de dados."""
    print(f"--- Iniciando processamento de XLSX: {caminho_arquivo} ---")
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo, read_only=True)
        sheet = workbook.active

        print("--- Planilha aberta com sucesso! ---")

        conexao = sqlite3.connect("database/banco_producao.db")
        cursor = conexao.cursor()
        print("--- Conexão com banco de dados estabelecida! ---")

        # Cria tabela caso nao exista
        cursor.execute("""
                       CREATE TABLE IF NOT EXISTS atividades_digitalizacao (
                           usuario TEXT,
                           ordem INTEGER,
                           cod_projeto INTEGER,
                           cod_lote INTEGER,
                           cod_pasta INTEGER,
                           projeto TEXT,
                           lote TEXT,
                           pasta TEXT,
                           fase TEXT,
                           fase_destino TEXT,
                           imgs_antes INTEGER,
                           imgs_depois INTEGER,
                           docs_antes INTEGER,
                           docs_depois INTEGER,
                           inicio TEXT,
                           termino TEXT,
                           tempo_gasto TEXT,
                           supervisor TEXT,
                           unidade TEXT,
                           envio TEXT
                       )
                   """)
        print("--- Tabela criada (ou já existente)! ---")

        iniciar_loading_screen(parent)  # Chama a loading screen
        print("Sucesso loading screen")

        num_linhas = sheet.max_row - 1  # Obtém o número de linhas a serem processadas (excluindo cabeçalho)
        linhas_processadas = 0

        # Itera sobre as linhas da planilha a partir da segunda linha (índice 1)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Cria uma lista com os dados da linha, pulando o campo "Computador" e "Local"
            usuario = row[0]
            # Ignora linhas em que o usuário é "Total"
            if usuario == "Total":
                continue

            row = ["" if cell is None else cell for cell in row]
            inicio = row[15].strftime("%Y-%m-%d %H:%M:%S")  # Formata a data
            termino = row[16].strftime("%Y-%m-%d %H:%M:%S")  # Formata a data
            tempo_gasto_str = row[17].strftime("%H:%M:%S") if row[17] is not None else ''
            data_hora_envio = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            dados = list(row[:1]) + list(row[2:17]) + [tempo_gasto_str, supervisor, unidade]
            dados.append(data_hora_envio)  # Adicione ao final da lista

            print(f"Dados a serem inseridos: {dados}")

            # Executa a consulta SQL para inserir os dados na tabela
            cursor.execute("""
                INSERT INTO atividades_digitalizacao (
                    usuario, ordem, cod_projeto, cod_lote, cod_pasta, projeto, lote, 
                    pasta, fase, fase_destino, imgs_antes, imgs_depois, docs_antes, 
                    docs_depois, inicio, termino, tempo_gasto, supervisor, unidade, envio
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, dados)
            print("--- Linha inserida com sucesso! ---")

            # --- Atualizar a Barra de Progresso ---
            linhas_processadas += 1
            progresso = int((linhas_processadas / num_linhas) * 100)
            parent.loading_screen.atualizar_progresso(progresso)  # Chama o método da loading screen

        conexao.commit()
        conexao.close()

        parent.loading_screen.finalizar_loading()  # Chama a função para finalizar a loading screen
        QMessageBox.information(None, "Sucesso",
                                f"Dados do arquivo '{caminho_arquivo}' inseridos no banco de dados com sucesso!")

        print(f"--- Dados do arquivo '{caminho_arquivo}' inseridos no banco de dados. ---")

    except Exception as e:
        print(f"Erro ao abrir ou processar o arquivo .xlsx: {e}")
    print(f"--- Fim do processamento de XLSX: {caminho_arquivo} ---")


def dicionario_txt(caminho_arquivo, supervisor, unidade, parent):
    """Processa um arquivo .txt e insere os dados no banco de dados."""
    print(f"--- Iniciando processamento de TXT: {caminho_arquivo} ---")
    try:
        conexao = sqlite3.connect("database/banco_producao.db")
        cursor = conexao.cursor()
        print("--- Conexão com banco de dados estabelecida! ---")

        # Cria tabela caso nao exista
        cursor.execute("""
                       CREATE TABLE IF NOT EXISTS atividades_digitalizacao (
                           usuario TEXT,
                           ordem INTEGER,
                           cod_projeto INTEGER,
                           cod_lote INTEGER,
                           cod_pasta INTEGER,
                           projeto TEXT,
                           lote TEXT,
                           pasta TEXT,
                           fase TEXT,
                           fase_destino TEXT,
                           imgs_antes INTEGER,
                           imgs_depois INTEGER,
                           docs_antes INTEGER,
                           docs_depois INTEGER,
                           inicio TEXT,
                           termino TEXT,
                           tempo_gasto TEXT,
                           supervisor TEXT,
                           unidade TEXT,
                           envio TEXT
                       )
                   """)
        print("--- Tabela criada (ou já existente)! ---")

        with open(caminho_arquivo, 'r', encoding='latin-1') as arquivo:
            num_linhas = sum(1 for linha in arquivo) - 1  # Conta as linhas do arquivo, excluindo o cabeçalho
            arquivo.seek(0)  # Volta para o inicio do arquivo
            next(arquivo)  # Pula o cabeçalho
            linhas_processadas = 0

            iniciar_loading_screen(parent)  # Chama a loading screen
            print("Sucesso loading screen")
            for linha in arquivo:
                # Ignora linhas que começam com "Total"
                if linha.startswith("Total"):
                    continue
                print(f"Processando linha: {linha.strip()}")
                # Remove o ponto e vírgula extra do final da linha, se existir
                if linha.endswith(";"):
                    linha = linha[:-1]

                dados = linha.strip().split(';')
                dados = ["" if dado is None else dado for dado in dados]
                del dados[1]  # Remove "Computador"
                del dados[17]  # Remove local

                # Ignora o último campo se ele for vazio ou tiver apenas espaços em branco
                if dados[-1].strip() == "":
                    dados = dados[:-1]

                dados.extend([supervisor, unidade])
                data_hora_envio = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                # Corrigindo o formato da data
                dados[14] = converter_para_data(dados[14], formato_entrada="%m/%d/%Y %H:%M:%S")
                dados[15] = converter_para_data(dados[15], formato_entrada="%m/%d/%Y %H:%M:%S")
                dados.append(data_hora_envio)  # Adicione ao final da lista
                print(f"Dados a serem inseridos: {dados}")
                cursor.execute("""
                    INSERT INTO atividades_digitalizacao (
                        usuario, ordem, cod_projeto, cod_lote, cod_pasta, projeto, lote, 
                        pasta, fase, fase_destino, imgs_antes, imgs_depois, docs_antes, 
                        docs_depois, inicio, termino, tempo_gasto, supervisor, unidade, envio
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, dados)
                print("--- Linha inserida com sucesso! ---")

                linhas_processadas += 1
                progresso = int((linhas_processadas / num_linhas) * 100)
                parent.loading_screen.atualizar_progresso(progresso)

        conexao.commit()
        conexao.close()

        parent.loading_screen.finalizar_loading()  # Chama a função para finalizar a loading screen
        QMessageBox.information(None, "Sucesso",
                                f"Dados do arquivo '{caminho_arquivo}' inseridos no banco de dados com sucesso!")

        print(f"--- Dados do arquivo '{caminho_arquivo}' inseridos no banco de dados. ---")

    except sqlite3.Error as e:
        print(f"Erro ao inserir dados no banco de dados: {e}")
    print(f"--- Fim do processamento de TXT: {caminho_arquivo} ---")


def validar_arquivo(caminho_arquivo, supervisor, unidade, parent):
    """Valida e processa o arquivo."""
    print(f"--- Iniciando validação do arquivo: {caminho_arquivo} ---")
    print(f"Supervisor: {supervisor}, Unidade: {unidade}")
    nome_arquivo, extensao = os.path.splitext(caminho_arquivo)
    if extensao.lower() == '.txt':
        # Validação do cabeçalho para arquivos .txt
        cabecalho_esperado = "Usuário;Computador;Ordem;Cod. Projeto;Cod. Lote;Cod. Pasta;Projeto;Lote;Pasta;Fase;Fase Destino;Imgs. Antes;Imgs. Depois;Docs. Antes;Docs. Depois;Início;Término;Tempo Gasto;Local;"
        with open(caminho_arquivo, 'r', encoding='latin-1') as arquivo:
            primeira_linha = arquivo.readline().strip()  # Lê a primeira linha
            # Aplica strip() a cada elemento do cabeçalho do arquivo
            cabecalho_arquivo = [campo.strip() for campo in primeira_linha.split(';')]
            # Aplica strip() a cada elemento do cabeçalho esperado
            cabecalho_esperado = [campo.strip() for campo in cabecalho_esperado.split(';')]
            if cabecalho_arquivo != cabecalho_esperado:
                print(f"Erro: Cabeçalho do arquivo .txt inválido!")
                return False
        print("--- Arquivo TXT válido! Chamando dicionario_txt... ---")
        dicionario_txt(caminho_arquivo, supervisor, unidade, parent)
        return True
    elif extensao.lower() == '.xlsx':
        try:
            workbook = openpyxl.load_workbook(caminho_arquivo, read_only=True)
            sheet = workbook.active  # Pega a planilha ativa

            # Validação do número de colunas
            if sheet.max_column != 19:
                print(
                    f"Erro: Número de colunas inválido no arquivo .xlsx! (Esperado: 19, Encontrado: {sheet.max_column})")
                return False

            # Validação dos nomes das colunas (cabeçalho)
            cabecalho_esperado = ["Usuário", "Computador", "Ordem", "Cod. Projeto", "Cod. Lote", "Cod. Pasta",
                                  "Projeto", "Lote", "Pasta", "Fase", "Fase Destino", "Imgs. Antes",
                                  "Imgs. Depois", "Docs. Antes", "Docs. Depois", "Início", "Término",
                                  "Tempo Gasto", "Local"]

            for i in range(1, sheet.max_column + 1):  # Percorre as colunas do cabeçalho
                if sheet.cell(row=1, column=i).value != cabecalho_esperado[i - 1]:
                    print(f"Erro: Nome da coluna inválido no arquivo .xlsx! (Coluna: {i})")
                    return False
        except Exception as e:
            print(f"Erro ao abrir ou processar o arquivo .xlsx: {e}")
            return False
        print("--- Arquivo XLSX válido! Chamando dicionario_xlsx... ---")
        dicionario_xlsx(caminho_arquivo, supervisor, unidade, parent)
        return True
    else:
        print("--- Extensão de arquivo inválida! ---")
        return False


def excluir_dados_banco():
    """Exclui todos os dados da tabela 'atividades_digitalizacao'."""
    try:
        conexao = sqlite3.connect("database/banco_producao.db")
        cursor = conexao.cursor()
        cursor.execute("DELETE FROM atividades_digitalizacao")
        conexao.commit()
        conexao.close()
        print("Todos os dados do banco de dados foram excluídos.")
    except sqlite3.Error as e:
        print(f"Erro ao excluir dados do banco de dados: {e}")


def converter_para_data(data_str, formato_entrada="%m/%d/%Y %H:%M:%S"):
    """Converte uma string de data para o formato padrão 'aaaa-mm-dd hh:mm:ss'.

    Args:
        data_str (str): A string da data a ser convertida.
        formato_entrada (str): O formato da data de entrada.

    Returns:
        str: A data no formato "aaaa-mm-dd hh:mm:ss", ou None se a conversão falhar.
    """
    print(f"--- Convertendo data: {data_str} ---")
    try:
        data_datetime = datetime.datetime.strptime(data_str, formato_entrada)
        print(f"Data convertida: {data_datetime}")
        return data_datetime.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError as e:
        print(f"Erro na conversão de data: {e}")
        return None

def iniciar_loading_screen(self):
    """Inicia a loading screen após o login bem-sucedido."""
    self.loading_screen = loading_interpretador.LoadingScreen()
    self.loading_screen.show()